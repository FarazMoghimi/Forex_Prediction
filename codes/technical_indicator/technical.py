from openpyxl import load_workbook, Workbook
import numpy as np
from sklearn import svm
from pyti.moving_average_convergence_divergence import moving_average_convergence_divergence as macd_func
from pyti.money_flow_index import money_flow_index as mfi_func
from pyti.relative_strength_index import relative_strength_index as rsi_func
from pyti.simple_moving_average import simple_moving_average as ma_func
from pyti.smoothed_moving_average import smoothed_moving_average as sma_func

wb2 = load_workbook('XAUUSD1440.xlsx', data_only=True)
ws = wb2.get_active_sheet()


close_data = np.array([i.value for i in ws['F']])[1:].astype(np.float).tolist()
open_data = np.array([i.value for i in ws['C']])[1:].astype(np.float).tolist()
high_data = fix_missing(np.array([i.value for i in ws['D']])[1:].astype(np.float).tolist())
low_data = fix_missing(np.array([i.value for i in ws['E']])[1:].astype(np.float).tolist())
y_predict_data = ['y_predict'] + y_predict(close_data, 5)
#close_data = ma_func(close_data, 4).tolist()
#high_data = ma_func(high_data, 4).tolist()
#low_data = ma_func(low_data, 4).tolist()
#close_data[0:4] = high_data[0:4] = low_data[0:4] = [0, 0, 0, 0]
volume_data = np.array([i.value for i in ws['G']])[1:].astype(np.float).tolist()
macd_data = ['macd'] + list(macd_func(close_data, short_period=12, long_period=26))
rsi14_data = ['rsi_14'] + [0 for _ in range(4)] + list(rsi_func(close_data[4:], period=14))
rsi9_data = ['rsi_9'] +  [0 for _ in range(4)] + list(rsi_func(close_data[4:], period=9))
ma21_data = ['ma_21'] + list(ma_func(close_data, period=21))
ma5_data = ['ma_5'] + list(ma_func(close_data, period=5))
sma5_s3_data = ['sma_5_shift3'] + [0, 0, 0] + list(sma_func(close_data, period=5))[:-3]
sma8_s5_data = ['sma_8_shift5'] + [0, 0, 0, 0, 0] + list(sma_func(close_data, period=8))[:-5]
mfi14_data = ['mfi_14'] + list(
    mfi_func(close_data=close_data, high_data=high_data, volume=volume_data, low_data=low_data, period=14))
rsi14_slope_data = ['rsi_14_slope'] + slope(rsi14_data[1:])
mfi14_bound_data = ['mfi_14_bound'] + mfi_bound(mfi14_data[1:])
mfi14_div = ['mfi_14_div'] + mfi_div(mfi14_data[1:], close_data)
macd14_rule_data = ['macd_14_rule'] + macd_rule(macd_data[1:], close_data)
ma21_rule = ['ma_21_rule'] + ma_rule(ma21_data[1:], close_data)
ma5_rule = ['ma_5_rule'] + ma_rule(ma5_data[1:], close_data)
rsi14_border = ['rsi_14_border'] + rsi_border(rsi14_data[1:])
rsi14_50_border = ['rsi_14_50_border'] + rsi_border(rsi14_data[1:], 60, 40)
rsi9_border = ['rsi_9_border'] + rsi_border(rsi9_data[1:])
ma5_slope_data = ['ma_5_slope'] + ma_slope(ma5_data[1:])
ma21_slope_data = ['ma_21_slope'] + ma_slope(ma21_data[1:])
k_percent_data = ['k%'] + k_percent(high_data, low_data, close_data)
d_percent_data = ['d%'] + d_percent(k_percent_data[1:])
hcl_data = ['hcl'] + hcl(close_data, high_data, low_data)
range_data = ['range'] + range_function(close_data, high_data, low_data)
min1_data = ['min1'] + min_local(close_data)
max1_data = ['max1'] + max_local(close_data)
ma9_data = ['ma_9'] + list(ma_func(close_data, period=9))
ma26_data = ['ma_26'] + list(ma_func(close_data, period=26))
ma9_ma26_cmp_data = ['ma_9_26_cmp'] + compare(ma9_data[1:], ma26_data[1:])
ma9_rule = ['ma_9_rule'] + ma_rule(ma9_data[1:], close_data)
ma26_rule = ['ma_26_rule'] + ma_rule(ma26_data[1:], close_data)
min4_data = ['min4'] + min_local(close_data, 4)
max4_data = ['max4'] + max_local(close_data, 4)
swing_data = ['swing'] + swing(max4_data[1:], min4_data[1:])
ma9_slope_data = ['ma_9_slope'] + ma_slope(ma9_data[1:])
ma26_slope_data = ['ma_26_slope'] + ma_slope(ma26_data[1:])
rsi14_step_data = ['rsi_14_step'] + step_func(rsi14_data[1:])
rsi9_step_data = ['rsi_9_step'] + step_func(rsi9_data[1:])
sma5_sma8_data = ['sma5_sma8'] + compare(sma5_s3_data, sma8_s5_data)
price1_slope_data = ['price_1_slope'] + slope(close_data, 1)
price2_slope_data = ['price_2_slope'] + slope(close_data, 2)
close_history_data = []
for i in range(1, 10):
    tmp = ['close_' + str(i)] + [close_data[j-i] for j in range(len(close_data))]
    close_history_data.append(tmp)

k_percent_rule_data = ['k%_rule'] + k_percent_rule(k_percent_data[1:])
k_percent_d_percent_data = ['k%_d%'] + k_percent_d_percent(d_percent_data[1:], k_percent_data[1:])


##### save
matrix = [[i.value for i in j] for j in ws]
wb_result = Workbook()
ws_raw = wb_result.create_sheet('raw')
ws_x = wb_result.create_sheet('x')
ws_y = wb_result.create_sheet('y')
ws_others = wb_result.create_sheet('others')

for count, value in enumerate(matrix):
    ws_raw.append(value)
    ws_others.append([k_percent_data[count],
                      d_percent_data[count],
                      rsi14_data[count],
                      rsi9_data[count],
                      ma21_data[count],
                      mfi14_data[count],
                      macd_data[count],
                      min1_data[count],
                      max1_data[count],
                      ma9_data[count],
                      ma26_data[count],
                      min4_data[count],
                      max4_data[count],
                      sma5_s3_data[count],
                      sma8_s5_data[count]])
    ws_x.append([ma21_rule[count],
                 rsi14_border[count],
                 rsi9_border[count],
                #rsi14_50_border[count],
                 ma5_rule[count],
                 ma5_slope_data[count],
                 ma21_slope_data[count],
                 ma9_slope_data[count],
                 ma26_slope_data[count],
                 rsi14_slope_data[count],
                 k_percent_rule_data[count],
                 k_percent_d_percent_data[count],
                 macd14_rule_data[count],
                 mfi14_bound_data[count],
                 mfi14_div[count],
                 #hcl_data[count],
                 #range_data[count],
                 #ma9_ma26_cmp_data[count],
                 #ma9_rule[count],
                 #ma26_rule[count],
                 #rsi14_step_data[count],
                 #rsi9_step_data[count],
                 #sma5_sma8_data[count],
                 #price1_slope_data[count],
                 #price2_slope_data[count],
                 # close_history_data[1][count],
                 # close_history_data[2][count],
                 # close_history_data[3][count],
                 # close_history_data[4][count],
                 # close_history_data[5][count],
                 # close_history_data[6][count],
                 # close_history_data[7][count],
                 #swing_data[count]
                 ])

    ws_y.append([y_predict_data[count]])

wb_result.save("new_technical.xlsx")
